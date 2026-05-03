"""Build a single ground-truth xlsx for the 62 subjects in kfold_splits_processed_new_mc.json.

Sources merged (in priority order for overlapping fields):
  1. data/splits/kfold_splits_processed_new_mc.json   -> canonical subject list + group label
  2. references/analysis/Anxiety Score Scale_Cleaned.xlsx -> HAMA/STAIS/STAIT items + sums (older cohort)
  3. references/synology_0331/.../rawdata/fNIRS_data20260323.xlsx -> scale + fNIRS file ids (newer cohort)
  4. references/analysis/Record Invitation.xlsx       -> demographics, diagnosis, measurement dates
  5. data_NIRx/Toolchain/filenamelist_20240305.csv    -> per-task file id (older cohort)
  6. data_NIRx/Toolchain/nirs_metadata.csv            -> per-task file id (newer cohort)

ID alias resolution: AH017..AH028 in kfold are recorded as EH017..EH028 in Record Invitation
(same physical subject, different institution prefix). We canonicalise to the kfold form.

Output: data/subjects_ground_truth.xlsx with sheets:
  - subjects     : 62-row summary, one row per subject, all key columns flattened
  - scales_items : item-level HAMA(_1..14)/STAIS(_1..20)/STAIT(_1..20) per subject
  - nirs_files   : long-format (subject, task, file_id) mapping
  - metadata     : provenance and generation details
"""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path("/home/user/jeffrymahbuubi/PROJECTS/2-fNIRS-Graph-Base-Method")

KFOLD_JSON = ROOT / "data/splits/kfold_splits_processed_new_mc.json"
ANX_XLSX = ROOT / "references/analysis/Anxiety Score Scale_Cleaned.xlsx"
RI_XLSX = ROOT / "references/analysis/Record Invitation.xlsx"
FNIRS_NEW_XLSX = (
    ROOT
    / "references/synology_0331/clinical data/rawdata/fNIRS_data20260323.xlsx"
)
DATA_NIRX = ROOT / "references/synology_0331/clinical data/rawdata/data_NIRx"
NIRS_META_CSV = DATA_NIRX / "Toolchain/nirs_metadata.csv"
FILENAMELIST_CSV = DATA_NIRX / "Toolchain/filenamelist_20240305.csv"

OUT_XLSX = ROOT / "data/subjects_ground_truth.xlsx"


# AH017..AH028 (kfold form) <-> EH017..EH028 (Record Invitation form)
AH_EH_ALIAS = {f"AH{i:03d}": f"EH{i:03d}" for i in range(17, 29)}
EH_AH_ALIAS = {v: k for k, v in AH_EH_ALIAS.items()}


def _norm(s):
    return s.strip() if isinstance(s, str) else s


def _date(d):
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d")
    return d


def load_kfold():
    with open(KFOLD_JSON) as f:
        data = json.load(f)
    return data["subjects"], data


def load_anxiety_scale():
    """Return {subject_id: row_dict} keyed by canonical id."""
    wb = openpyxl.load_workbook(ANX_XLSX, data_only=True)
    ws = wb["Sheet1"]
    headers = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        sid = _norm(row[0])
        if not isinstance(sid, str):
            continue
        out[sid] = dict(zip(headers, row))
    return out, headers


def load_record_invitation():
    """Return {canonical_subject_id: row_dict}. Applies EH->AH alias."""
    wb = openpyxl.load_workbook(RI_XLSX, data_only=True)
    ws = wb["Sheet1"]
    headers = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        raw_id = _norm(row[1])
        if not isinstance(raw_id, str):
            continue
        canonical = EH_AH_ALIAS.get(raw_id, raw_id)
        d = dict(zip(headers, row))
        d["_raw_id"] = raw_id  # preserve original id for traceability
        out[canonical] = d
    return out, headers


def load_fnirs_new():
    """Return {subject_id: {scale: row_dict, fnirs: row_dict, task_order: row_dict}}."""
    wb = openpyxl.load_workbook(FNIRS_NEW_XLSX, data_only=True)
    out = defaultdict(dict)

    # 量表分數 sheet (scale scores)
    ws = wb["量表分數"]
    hdr_scale = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        sid = _norm(row[0])
        if isinstance(sid, str):
            out[sid]["scale"] = dict(zip(hdr_scale, row))

    # fNIRS sheet (file mapping)
    ws = wb["fNIRS"]
    hdr_fn = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        sid = _norm(row[0])
        if isinstance(sid, str):
            out[sid]["fnirs"] = dict(zip(hdr_fn, row))

    # task order
    ws = wb["task order"]
    hdr_to = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        sid = _norm(row[0])
        if isinstance(sid, str):
            out[sid]["task_order"] = dict(zip(hdr_to, row))

    return dict(out), hdr_scale, hdr_fn, hdr_to


def load_nirs_metadata():
    """Return {subject_id: {file_id, group, task_count, task_1, task_2, notes}}."""
    out = {}
    with open(NIRS_META_CSV) as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            sid = _norm(row.get("subject"))
            if isinstance(sid, str):
                # one row per file; keep all rows under list
                out.setdefault(sid, []).append(row)
    return out


def load_filenamelist():
    """Return {subject_id: {task: file_id}}."""
    out = defaultdict(dict)
    with open(FILENAMELIST_CSV) as fh:
        reader = csv.reader(fh)
        for row in reader:
            if len(row) < 3 or not row[0]:
                continue
            sid = row[0].strip().lstrip("﻿")
            task = (row[1] or "").strip()
            fid = (row[2] or "").strip()
            if task:
                out[sid][task] = fid
    return dict(out)


def main():
    kfold_subjects, kfold_meta = load_kfold()
    anx, anx_headers = load_anxiety_scale()
    ri, ri_headers = load_record_invitation()
    fnirs_new, fn_scale_hdr, fn_fn_hdr, fn_to_hdr = load_fnirs_new()
    nirs_meta = load_nirs_metadata()
    filenamelist = load_filenamelist()

    # Build "subjects" main sheet
    subject_rows = []
    coverage = {"anx": 0, "ri": 0, "fnirs_new": 0, "filenamelist": 0, "nirs_meta": 0}

    for sid in sorted(kfold_subjects.keys()):
        group = kfold_subjects[sid]

        anx_row = anx.get(sid, {})
        ri_row = ri.get(sid, {})
        fnirs_row = fnirs_new.get(sid, {})
        scale_new = fnirs_row.get("scale", {})
        fnirs_filemap = fnirs_row.get("fnirs", {})
        task_order_new = fnirs_row.get("task_order", {})

        # Source flags
        src_anx = "Y" if anx_row else ""
        src_ri = "Y" if ri_row else ""
        src_fnirs_new = "Y" if scale_new or fnirs_filemap else ""

        # Demographics: prefer Record Invitation
        birth = _date(ri_row.get("生日"))
        age = ri_row.get("年齡")
        gender = ri_row.get("性別")
        education = ri_row.get("教育程度")
        marital = ri_row.get("婚姻狀態")
        occupation = ri_row.get("職業")
        onset_year = ri_row.get("發病年份")
        diagnosis = ri_row.get("診斷")
        treatment = ri_row.get("治療種類")
        medication = ri_row.get("服用藥物")
        first_meas = _date(ri_row.get("第一次量測日期"))
        second_meas_invite = ri_row.get("第二次量測邀請紀錄")
        second_meas_done = _date(ri_row.get("第二次測量"))
        ri_raw_id = ri_row.get("_raw_id", "")
        alias_used = ri_raw_id if ri_raw_id and ri_raw_id != sid else ""

        # Scale sums: prefer Anxiety Score Scale (older cohort), fall back to fNIRS_data20260323
        if anx_row:
            hama_sum = anx_row.get("HAMA_sum")
            stais_sum = anx_row.get("STAIS_sum")
            stait_sum = anx_row.get("STAIT_sum")
        else:
            hama_sum = scale_new.get("HAM_sum")
            stais_sum = scale_new.get("STAIS_sum")
            stait_sum = scale_new.get("STAIT_sum")
        scale_source = "Anxiety Score Scale_Cleaned" if anx_row else (
            "fNIRS_data20260323" if scale_new else ""
        )

        # Task order: prefer the text-format value in the fNIRS sheet
        # (the dedicated "task order" sheet stores it as 3/4 -> Excel mis-parses as date).
        task_order = fnirs_filemap.get(
            "功能性任務順序(VF:1 / 1backWM:2 / SS:3 / GNG:4 )"
        )
        if task_order is None:
            raw = task_order_new.get(
                "功能性任務順序(VF:1 / 1backWM:2 / SS:3 / GNG:4 )"
            )
            if isinstance(raw, datetime):
                # Excel parsed e.g. "3/4" -> 2025-03-04: recover month/day as task codes
                task_order = f"{raw.month}/{raw.day}"
            else:
                task_order = raw

        # fNIRS file mapping
        fl = filenamelist.get(sid, {})
        if fl:
            file_vf = fl.get("VF", "")
            file_1bw = fl.get("1backWM", "")
            file_ss = fl.get("SS", "")
            file_gng = fl.get("GNG", "")
            file_ap = fl.get("AP", "")
            file_rp = fl.get("RP", "")
            file_source = "filenamelist_20240305"
        else:
            file_vf = file_1bw = file_ss = file_gng = file_ap = file_rp = ""
            file_source = ""

        # Newer cohort: aggregate from nirs_metadata + fNIRS sheet
        nm_rows = nirs_meta.get(sid, [])
        nm_file_ids = "; ".join(r.get("file_id", "") for r in nm_rows)
        nm_tasks = "; ".join(
            "+".join(filter(None, [r.get("task_1", ""), r.get("task_2", "")]))
            for r in nm_rows
        )
        nm_notes = " | ".join(r.get("notes", "") for r in nm_rows if r.get("notes"))
        if not file_source and nm_rows:
            file_source = "nirs_metadata.csv"
        # Fall back: per-task split from nirs_metadata
        if not (file_vf or file_1bw or file_ss or file_gng) and nm_rows:
            for r in nm_rows:
                fid = r.get("file_id", "")
                for t_col in ("task_1", "task_2"):
                    t = r.get(t_col, "")
                    if t == "VF" and not file_vf:
                        file_vf = fid
                    elif t == "1backWM" and not file_1bw:
                        file_1bw = fid
                    elif t == "SS" and not file_ss:
                        file_ss = fid
                    elif t == "GNG" and not file_gng:
                        file_gng = fid

        fnirs_other = fnirs_filemap.get("檔案編號", "")
        fnirs_note = fnirs_filemap.get("其他事項補充", "")

        if anx_row:
            coverage["anx"] += 1
        if ri_row:
            coverage["ri"] += 1
        if scale_new or fnirs_filemap:
            coverage["fnirs_new"] += 1
        if fl:
            coverage["filenamelist"] += 1
        if nm_rows:
            coverage["nirs_meta"] += 1

        # demographics_missing flag: Y if Record Invitation row is absent
        # (these subjects have no birth_date/age/gender/education/etc.)
        demo_missing = "Y" if not ri_row else ""

        subject_rows.append({
            "subject_id": sid,
            "group": group,
            "alias_id_in_record_invitation": alias_used,
            "demographics_missing": demo_missing,
            "src_anxiety_scale": src_anx,
            "src_record_invitation": src_ri,
            "src_fnirs_data20260323": src_fnirs_new,
            "birth_date": birth,
            "age": age,
            "gender": gender,
            "education": education,
            "marital_status": marital,
            "occupation": occupation,
            "onset_year": onset_year,
            "diagnosis": diagnosis,
            "treatment": treatment,
            "medication": medication,
            "first_measurement_date": first_meas,
            "second_measurement_invite_note": second_meas_invite,
            "second_measurement_date": second_meas_done,
            "HAMA_sum": hama_sum,
            "STAIS_sum": stais_sum,
            "STAIT_sum": stait_sum,
            "scale_source": scale_source,
            "task_order_code": task_order,
            "fnirs_file_VF": file_vf,
            "fnirs_file_1backWM": file_1bw,
            "fnirs_file_SS": file_ss,
            "fnirs_file_GNG": file_gng,
            "fnirs_file_AP": file_ap,
            "fnirs_file_RP": file_rp,
            "fnirs_file_source": file_source,
            "fnirs_aggregated_file_ids": nm_file_ids or fnirs_other or "",
            "fnirs_aggregated_tasks": nm_tasks,
            "fnirs_notes": (nm_notes or fnirs_note or ""),
        })

    # Build "scales_items" sheet (long-form item-level for HAMA/STAIS/STAIT)
    item_rows = []
    hama_keys = [f"HAMA_{i}" for i in range(1, 15)]
    stais_keys = [f"STAIS_{i}" for i in range(1, 21)]
    stait_keys = [f"STAIT_{i}" for i in range(1, 21)]

    for sid in sorted(kfold_subjects.keys()):
        anx_row = anx.get(sid, {})
        scale_new = fnirs_new.get(sid, {}).get("scale", {})

        if anx_row:
            row = {"subject_id": sid, "source": "Anxiety Score Scale_Cleaned"}
            for k in hama_keys + ["HAMA_sum"] + stais_keys + ["STAIS_sum"] + stait_keys + ["STAIT_sum"]:
                row[k] = anx_row.get(k)
            item_rows.append(row)
        elif scale_new:
            row = {"subject_id": sid, "source": "fNIRS_data20260323"}
            # fNIRS_data20260323 uses "HAM_*" instead of "HAMA_*"
            for i, k in enumerate(hama_keys, start=1):
                row[k] = scale_new.get(f"HAM_{i}")
            row["HAMA_sum"] = scale_new.get("HAM_sum")
            for k in stais_keys + ["STAIS_sum"] + stait_keys + ["STAIT_sum"]:
                row[k] = scale_new.get(k)
            item_rows.append(row)
        else:
            item_rows.append({"subject_id": sid, "source": "MISSING"})

    # Build "nirs_files" sheet (long-form per-task file mapping)
    file_rows = []
    for sid in sorted(kfold_subjects.keys()):
        fl = filenamelist.get(sid, {})
        for task, fid in fl.items():
            if fid:
                file_rows.append({
                    "subject_id": sid,
                    "task": task,
                    "file_id": fid,
                    "source": "filenamelist_20240305.csv",
                })
        for r in nirs_meta.get(sid, []):
            for t_col in ("task_1", "task_2"):
                t = r.get(t_col, "")
                if t:
                    file_rows.append({
                        "subject_id": sid,
                        "task": t,
                        "file_id": r.get("file_id", ""),
                        "source": "nirs_metadata.csv",
                    })

    # Write the workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # subjects sheet
    ws = wb.create_sheet("subjects")
    if subject_rows:
        cols = list(subject_rows[0].keys())
        ws.append(cols)
        for r in subject_rows:
            ws.append([r.get(c) for c in cols])

    # scales_items sheet
    ws = wb.create_sheet("scales_items")
    item_cols = ["subject_id", "source"] + hama_keys + ["HAMA_sum"] + stais_keys + ["STAIS_sum"] + stait_keys + ["STAIT_sum"]
    ws.append(item_cols)
    for r in item_rows:
        ws.append([r.get(c) for c in item_cols])

    # nirs_files sheet
    ws = wb.create_sheet("nirs_files")
    file_cols = ["subject_id", "task", "file_id", "source"]
    ws.append(file_cols)
    for r in file_rows:
        ws.append([r.get(c) for c in file_cols])

    # metadata sheet
    ws = wb.create_sheet("metadata")
    ws.append(["key", "value"])
    md = [
        ("generated_at", datetime.now().isoformat(timespec="seconds")),
        ("script", "scripts/build_subject_ground_truth.py"),
        ("reference_split", str(KFOLD_JSON.relative_to(ROOT))),
        ("reference_split_total_subjects", kfold_meta.get("total_subjects")),
        ("reference_split_class_distribution", json.dumps(kfold_meta.get("class_distribution"))),
        ("source_anxiety_scale", str(ANX_XLSX.relative_to(ROOT))),
        ("source_record_invitation", str(RI_XLSX.relative_to(ROOT))),
        ("source_fnirs_data20260323", str(FNIRS_NEW_XLSX.relative_to(ROOT))),
        ("source_filenamelist", str(FILENAMELIST_CSV.relative_to(ROOT))),
        ("source_nirs_metadata", str(NIRS_META_CSV.relative_to(ROOT))),
        ("alias_rule", "AH017..AH028 in kfold == EH017..EH028 in Record Invitation (canonical=AH form)"),
        ("scale_priority", "Anxiety Score Scale_Cleaned > fNIRS_data20260323"),
        ("file_priority", "filenamelist_20240305.csv > nirs_metadata.csv"),
        ("excluded_pii", "姓名 (name) deliberately omitted from output"),
        ("coverage_anxiety_scale", f"{coverage['anx']}/62"),
        ("coverage_record_invitation", f"{coverage['ri']}/62"),
        ("coverage_fnirs_data20260323", f"{coverage['fnirs_new']}/62"),
        ("coverage_filenamelist", f"{coverage['filenamelist']}/62"),
        ("coverage_nirs_metadata", f"{coverage['nirs_meta']}/62"),
    ]
    for k, v in md:
        ws.append([k, v])

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    print(f"Wrote: {OUT_XLSX}")
    print(f"Subjects: {len(subject_rows)}")
    print("Source coverage:")
    for k, v in coverage.items():
        print(f"  {k}: {v}/62")


if __name__ == "__main__":
    main()
